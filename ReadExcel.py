import openpyxl

path = "C:\PycharmProjects\ExcelFiles\Excel1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active  #if your excel file has only one sheet
# sheet= workbook.get_sheet_by_name("Sheet1")

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r, column=c).value,end="  ")

    print()

